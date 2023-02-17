import socket
from tkinter import *

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Open a window
window = Tk()
window.title("Chip Counter")
window.geometry("250x270")

# Label
label = Label(window, text="Input Amount:", font=("Arial", 10))
label.grid(row=1, column=0)

Connect_b = Label(window, text="Connect to WizFi360", font=("Arial", 10))
Connect_b.grid(row=0, column=0)

label_excel = Label(window, text="Excel", font=("Arial", 10))
label_excel.grid(row=4, column=0)

Product_name = Label(window, text="Product Name:", font=("Arial", 10))
Product_name.grid(row=5, column=0)

Sheet_name = Label(window, text="Sheet Name:", font=("Arial", 10))
Sheet_name.grid(row=6, column=0)

label_file = Label(window, text="Name of the file:", font=("Arial", 10))
label_file.grid(row=7, column=0)

Ex_status = Label(window)
Ex_status.grid(row=8, column=0, sticky=EW)

# Entry
A_amount = Entry(window, width=10, font=("Arial", 10))
A_amount.grid(row=1, column=1)

PN_input = Entry(window, width=10, font=("Arial", 10))
PN_input.grid(row=5, column=1)

sheet_input = Entry(window, width=10, font=("Arial", 10))
sheet_input.grid(row=6, column=1)

file_name = Entry(window, width=10, font=("Arial", 10))
file_name.grid(row=7, column=1)

# Display
D_amount = Text(window, width=10, height=5, font=("Arial", 10))
D_amount.grid(row=2, column=0)
D_amount.configure(state=DISABLED)


def clear() -> NONE:
    D_amount.configure(state=NORMAL)
    D_amount.delete("1.0", END)
    D_amount.configure(state=DISABLED)


class Wifi360Network:
    HOST = "10.0.1.113"  # Standard loopback interface address (localhost)
    PORT = 5000  # Port to listen on (non-privileged ports are > 1023)
    status = 0  # 0 = not connected, 1 = connected
    data = None
    popup = None

    def __init__(self):
        self.s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.total_count = None
        self.file_id = None
        self.sheet_id = None
        self.product_id = None
        self.wb = None
        self.ws = None

    def change(self):
        chip = A_amount.get()
        self.s.sendall(chip.encode())

    def create(self):
        if self.status == 0:
            self.s.connect((self.HOST, self.PORT))
            Network.configure(text='Connected')
            self.status = 1
        elif self.status == 1:
            self.s.close()
            Network.configure(text='Connect')
            self.status = 0
            self.s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    def receive(self):
        print("before receive data")
        self.data = self.s.recv(1024)
        print(self.data)
        if self.data:
            D_amount.configure(state=NORMAL)
            D_amount.insert(END, "Counted Amount:\r\n")
            D_amount.insert(END, str(self.data.decode()))
            D_amount.configure(state=DISABLED)
            print("Collected:" + str(self.data))
        else:
            D_amount.configure(state=NORMAL)
            D_amount.insert(END, "No data")
            D_amount.configure(state=DISABLED)

    def count(self):
        count_mode = "CV"
        self.s.sendall(count_mode.encode())

    def find_cell(self):
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row + 1, max_col=1):
            for cell in row:
                cell_r = cell.row
                if cell.value == self.product_id:
                    cell_c = get_column_letter(cell.column + 1)
                    msg = "found"
                    return msg, cell_r, cell_c
                else:
                    cell_c = get_column_letter(cell.column)
                    if self.ws[str(cell_c) + str(cell_r + 1)].value is None:
                        msg = "not found"
                        return msg, cell_r, cell_c

    def create_excel(self):
        self.product_id = PN_input.get()
        self.sheet_id = sheet_input.get()
        self.file_id = file_name.get()
        self.total_count = int(self.data)
        try:
            self.wb = openpyxl.load_workbook(self.file_id)
            self.ws = self.wb[self.sheet_id]
            msg, cell_r, cell_c = self.find_cell()
            if msg == "not found":
                msg = "Part No. not found, Do you want to add this part No.?"
                self.pop(msg, cell_r, cell_c)
            elif msg == "found":
                if self.ws[str(cell_c) + str(cell_r)].value is None:
                    self.ws[str(cell_c) + str(cell_r)] = self.total_count
                    msg = "Saved"
                else:
                    self.ws[str(cell_c) + str(cell_r)] = self.total_count + int(
                        self.ws[str(cell_c) + str(cell_r)].value)
                    msg = "Added"
                Ex_status.config(text=msg, font=("Arial", 10), fg="#F5190B")
            self.wb.save(self.file_id)
        except FileNotFoundError:
            msg = "No such file, Do you want to create this file?"
            self.pop(msg)

    def pop(self, msg: str, cell_r=None, cell_c=None):
        self.popup = Toplevel(window)
        self.popup.title("Error")
        self.popup.geometry("350x100")
        e_label = Label(self.popup, text=msg, font=("Arial", 10), padx=20)
        e_label.grid(row=0, column=0)
        y_button = Button(self.popup, text="Yes", command=lambda: self.choice("YES", msg, cell_r, cell_c),
                          font=("Arial", 10))
        y_button.grid(row=1, column=0, sticky=EW)
        n_button = Button(self.popup, text="No", command=lambda: self.choice("NO", msg, cell_r, cell_c),
                          font=("Arial", 10))
        n_button.grid(row=2, column=0, sticky=EW)

    def choice(self, option: str, msg: str, cell_r=None, cell_c=None):
        self.popup.destroy()
        if option == "YES":
            if msg == "Part No. not found, Do you want to add this part No.?":
                self.ws[str(cell_c) + str(cell_r + 1)].value = self.product_id
                cell_c = chr(ord(cell_c) + 1)
                self.ws[str(cell_c) + str(cell_r+1)].value = self.total_count
                Ex_status.config(text="Added new item", font=("Arial", 10), fg="#F5190B")
                self.wb.save(self.file_id)
            else:
                self.wb = Workbook()
                self.ws = self.wb.active
                self.ws.title = self.sheet_id
                self.ws['A1'] = "Product Name"
                self.ws['B1'] = "QTY"
                self.ws['A2'] = self.product_id
                self.ws['B2'] = self.total_count
                self.wb.save(self.file_id)
                Ex_status.config(text="Created a New file", font=("Arial", 10), fg="#F5190B")
        else:
            Ex_status.config(text="Check file location", font=("Arial", 10), fg="#F5190B")


wifi = Wifi360Network()

# Buttons
Network = Button(window, text="Connect", command=wifi.create, font=("Arial", 10))
Network.grid(row=0, column=1)

button = Button(window, text="Send", command=wifi.change, font=("Arial", 10))
button.grid(row=1, column=2)

Update = Button(window, text="Update", command=wifi.receive, font=("Arial", 10))
Update.grid(row=2, column=1)

Count_value = Button(window, text="Count", command=wifi.count, font=("Arial", 10))
Count_value.grid(row=0, column=2)

Clear = Button(window, text="Clear", command=clear, font=("Arial", 10))
Clear.grid(row=2, column=2)

Ex_save = Button(window, text="Save", command=wifi.create_excel, font=("Arial", 10))
Ex_save.grid(row=7, column=2)

window.mainloop()
